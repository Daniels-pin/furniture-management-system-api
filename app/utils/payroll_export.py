"""Month-end payroll export — presentation layer over existing payroll calculations."""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date as date_type, datetime
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import models
from app.constants import APP_NAME
from app.utils.timezone import now_lagos, to_lagos

_MONTH_NAMES = (
    "",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
)


def _decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def _format_day_label(d: date_type) -> str:
    return f"{_MONTH_NAMES[d.month]} {d.day}"


def _format_generated_at(dt: datetime) -> tuple[str, str]:
    local = to_lagos(dt)
    date_part = f"{local.day} {_MONTH_NAMES[local.month]} {local.year}"
    hour = local.hour % 12 or 12
    minute = f"{local.minute:02d}"
    am_pm = "AM" if local.hour < 12 else "PM"
    time_part = f"{hour}:{minute} {am_pm}"
    return date_part, time_part


def _payment_status_label(status: str | None) -> str:
    return "Paid" if status == "paid" else "Pending"


def _month_status_label(status: str | None) -> str:
    return "Paid" if status == "paid" else "Pending Payment"


@dataclass
class PayrollExportLineItem:
    label: str
    amount: Decimal
    date_label: Optional[str] = None


@dataclass
class PayrollExportEmployeeRow:
    employee_id: int
    name: str
    department: str
    bank_name: str
    account_name: str
    account_number: str
    salary_month: str
    base_salary: Decimal
    bonuses_total: Decimal
    increments_total: Decimal
    lateness_deduction: Decimal
    absence_deduction: Decimal
    early_sign_out_deduction: Decimal
    manual_deductions: Decimal
    other_adjustments: Decimal
    total_deductions: Decimal
    final_payable: Decimal
    payment_status: str
    bonus_lines: list[PayrollExportLineItem] = field(default_factory=list)
    increment_lines: list[PayrollExportLineItem] = field(default_factory=list)
    lateness_lines: list[PayrollExportLineItem] = field(default_factory=list)
    absence_lines: list[PayrollExportLineItem] = field(default_factory=list)
    early_sign_out_lines: list[PayrollExportLineItem] = field(default_factory=list)
    manual_deduction_lines: list[PayrollExportLineItem] = field(default_factory=list)


@dataclass
class PayrollExportSummary:
    employee_count: int
    total_base_salaries: Decimal
    total_bonuses: Decimal
    total_increments: Decimal
    total_deductions: Decimal
    grand_total_payable: Decimal


@dataclass
class PayrollExportPayload:
    company_name: str
    payroll_month: str
    payroll_status: str
    generated_by: str
    generated_date: str
    generated_time: str
    period_id: int
    period_year: int
    period_month: int
    employees: list[PayrollExportEmployeeRow]
    summary: PayrollExportSummary


def _batch_lateness_lines(db: Session, period_id: int, emp_ids: list[int]) -> dict[int, list[PayrollExportLineItem]]:
    if not emp_ids:
        return {}
    rows = (
        db.query(
            models.EmployeeLatenessEntry.employee_id,
            models.EmployeeLatenessEntry.deduction_amount_naira,
            models.EmployeeAttendanceEntry.attendance_date,
            models.EmployeeLatenessEntry.created_at,
        )
        .outerjoin(
            models.EmployeeAttendanceEntry,
            models.EmployeeLatenessEntry.attendance_id == models.EmployeeAttendanceEntry.id,
        )
        .filter(
            models.EmployeeLatenessEntry.period_id == period_id,
            models.EmployeeLatenessEntry.employee_id.in_(emp_ids),
            models.EmployeeLatenessEntry.voided_at.is_(None),
        )
        .order_by(models.EmployeeLatenessEntry.id)
        .all()
    )
    out: dict[int, list[PayrollExportLineItem]] = {eid: [] for eid in emp_ids}
    for eid, amount, att_date, created_at in rows:
        eid_int = int(eid)
        if att_date is not None:
            date_label = _format_day_label(att_date)
        elif created_at is not None:
            date_label = _format_day_label(to_lagos(created_at).date())
        else:
            date_label = None
        out.setdefault(eid_int, []).append(
            PayrollExportLineItem(label="Late Coming", amount=_decimal(amount), date_label=date_label)
        )
    return out


def _batch_absence_lines(db: Session, period_id: int, emp_ids: list[int]) -> dict[int, list[PayrollExportLineItem]]:
    if not emp_ids:
        return {}
    rows = (
        db.query(
            models.EmployeeAbsenceEntry.employee_id,
            models.EmployeeAbsenceEntry.deduction_amount_naira,
            models.EmployeeAbsenceEntry.absence_date,
        )
        .filter(
            models.EmployeeAbsenceEntry.period_id == period_id,
            models.EmployeeAbsenceEntry.employee_id.in_(emp_ids),
            models.EmployeeAbsenceEntry.voided_at.is_(None),
        )
        .order_by(models.EmployeeAbsenceEntry.absence_date, models.EmployeeAbsenceEntry.id)
        .all()
    )
    out: dict[int, list[PayrollExportLineItem]] = {eid: [] for eid in emp_ids}
    for eid, amount, absence_date in rows:
        eid_int = int(eid)
        date_label = _format_day_label(absence_date) if absence_date else None
        out.setdefault(eid_int, []).append(
            PayrollExportLineItem(label="Absence", amount=_decimal(amount), date_label=date_label)
        )
    return out


def _batch_early_sign_out_lines(db: Session, period_id: int, emp_ids: list[int]) -> dict[int, list[PayrollExportLineItem]]:
    if not emp_ids:
        return {}
    rows = (
        db.query(
            models.EmployeeEarlySignOutEntry.employee_id,
            models.EmployeeEarlySignOutEntry.deduction_amount_naira,
            models.EmployeeAttendanceEntry.attendance_date,
            models.EmployeeEarlySignOutEntry.created_at,
        )
        .outerjoin(
            models.EmployeeAttendanceEntry,
            models.EmployeeEarlySignOutEntry.attendance_id == models.EmployeeAttendanceEntry.id,
        )
        .filter(
            models.EmployeeEarlySignOutEntry.period_id == period_id,
            models.EmployeeEarlySignOutEntry.employee_id.in_(emp_ids),
            models.EmployeeEarlySignOutEntry.voided_at.is_(None),
        )
        .order_by(models.EmployeeEarlySignOutEntry.id)
        .all()
    )
    out: dict[int, list[PayrollExportLineItem]] = {eid: [] for eid in emp_ids}
    for eid, amount, att_date, created_at in rows:
        eid_int = int(eid)
        if att_date is not None:
            date_label = _format_day_label(att_date)
        elif created_at is not None:
            date_label = _format_day_label(to_lagos(created_at).date())
        else:
            date_label = None
        out.setdefault(eid_int, []).append(
            PayrollExportLineItem(label="Early Sign-Out", amount=_decimal(amount), date_label=date_label)
        )
    return out


def _batch_adjustment_lines(
    db: Session, period_id: int, emp_ids: list[int]
) -> dict[int, dict[str, list[PayrollExportLineItem]]]:
    if not emp_ids:
        return {}
    rows = (
        db.query(models.EmployeePayrollAdjustment)
        .filter(
            models.EmployeePayrollAdjustment.period_id == period_id,
            models.EmployeePayrollAdjustment.employee_id.in_(emp_ids),
            models.EmployeePayrollAdjustment.voided_at.is_(None),
        )
        .order_by(models.EmployeePayrollAdjustment.created_at, models.EmployeePayrollAdjustment.id)
        .all()
    )
    out: dict[int, dict[str, list[PayrollExportLineItem]]] = {
        eid: {"bonus": [], "deduction": [], "increment": []} for eid in emp_ids
    }
    for adj in rows:
        eid = int(adj.employee_id)
        item = PayrollExportLineItem(label=(adj.reason or adj.adjustment_type).strip(), amount=_decimal(adj.amount))
        bucket = out.setdefault(eid, {"bonus": [], "deduction": [], "increment": []})
        adj_type = (adj.adjustment_type or "").strip()
        if adj_type in bucket:
            bucket[adj_type].append(item)
    return out


def build_payroll_export_payload(db: Session, period: models.SalaryPeriod, *, generated_by: str) -> PayrollExportPayload:
    """Build export payload using the same payroll list context as list/summary endpoints."""
    from app.routes.employees import _PayrollListContext

    ctx = _PayrollListContext.build(db, period)
    emps = sorted(ctx.emps, key=lambda e: (e.full_name or "").lower())
    emp_ids = [e.id for e in emps]

    lateness_lines = _batch_lateness_lines(db, period.id, emp_ids)
    absence_lines = _batch_absence_lines(db, period.id, emp_ids)
    early_lines = _batch_early_sign_out_lines(db, period.id, emp_ids)
    adj_lines = _batch_adjustment_lines(db, period.id, emp_ids)

    generated_dt = now_lagos()
    gen_date, gen_time = _format_generated_at(generated_dt)

    rows: list[PayrollExportEmployeeRow] = []
    total_base = Decimal("0")
    total_bon = Decimal("0")
    total_inc = Decimal("0")
    total_ded = Decimal("0")
    total_net = Decimal("0")

    for emp in emps:
        pt = ctx.pen_map.get(emp.id, Decimal("0"))
        bt = ctx.bon_map.get(emp.id, Decimal("0"))
        it = ctx.inc_map.get(emp.id, Decimal("0"))
        payroll = ctx.payroll_map.get(emp.id)
        salary = ctx.salary_for(emp, pt, bt, payroll, it)

        if payroll is not None and payroll.period_base_salary is not None:
            base_display = _decimal(payroll.period_base_salary)
        else:
            base_display = _decimal(emp.base_salary)

        dept = ""
        loc = getattr(emp, "work_location", None)
        if loc is not None and getattr(loc, "name", None):
            dept = str(loc.name)

        pr_status = payroll.payment_status if payroll else "unpaid"
        emp_adj = adj_lines.get(emp.id, {"bonus": [], "deduction": [], "increment": []})

        row = PayrollExportEmployeeRow(
            employee_id=emp.id,
            name=emp.full_name or "",
            department=dept,
            bank_name=getattr(emp, "bank_name", None) or "",
            account_name=emp.full_name or "",
            account_number=emp.account_number or "",
            salary_month=period.label,
            base_salary=base_display,
            bonuses_total=salary.bonuses_total,
            increments_total=salary.increments_total,
            lateness_deduction=salary.lateness_deduction,
            absence_deduction=salary.absence_deduction,
            early_sign_out_deduction=salary.early_sign_out_deduction,
            manual_deductions=salary.penalties_entries_total,
            other_adjustments=Decimal("0"),
            total_deductions=salary.total_deductions,
            final_payable=salary.final_payable,
            payment_status=_payment_status_label(pr_status),
            bonus_lines=emp_adj.get("bonus", []),
            increment_lines=emp_adj.get("increment", []),
            lateness_lines=lateness_lines.get(emp.id, []),
            absence_lines=absence_lines.get(emp.id, []),
            early_sign_out_lines=early_lines.get(emp.id, []),
            manual_deduction_lines=emp_adj.get("deduction", []),
        )
        rows.append(row)
        total_base += base_display
        total_bon += salary.bonuses_total
        total_inc += salary.increments_total
        total_ded += salary.total_deductions
        total_net += salary.final_payable

    summary = PayrollExportSummary(
        employee_count=len(rows),
        total_base_salaries=total_base,
        total_bonuses=total_bon,
        total_increments=total_inc,
        total_deductions=total_ded,
        grand_total_payable=total_net,
    )

    return PayrollExportPayload(
        company_name=APP_NAME,
        payroll_month=period.label,
        payroll_status=_month_status_label(getattr(period, "month_payment_status", None)),
        generated_by=generated_by or "Admin",
        generated_date=gen_date,
        generated_time=gen_time,
        period_id=period.id,
        period_year=period.year,
        period_month=period.month,
        employees=rows,
        summary=summary,
    )


def _money_cell_value(amount: Decimal) -> float:
    return float(amount)


def _apply_header_style(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="111111")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def payroll_export_xlsx_bytes(payload: PayrollExportPayload) -> bytes:
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Payroll Summary"

    ws["A1"] = payload.company_name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{payload.payroll_month} Payroll"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = f"Payroll Status: {payload.payroll_status}"
    ws["A4"] = f"Generated by: {payload.generated_by}"
    ws["A5"] = f"Generated: {payload.generated_date}  {payload.generated_time}"

    headers = [
        "Name",
        "Employee ID",
        "Department",
        "Bank",
        "Account Name",
        "Account Number",
        "Salary Month",
        "Base Salary",
        "Bonuses",
        "Salary Increments",
        "Late Deduction",
        "Absence Deduction",
        "Early Sign-Out Deduction",
        "Manual Deduction",
        "Other Adjustments",
        "Total Deductions",
        "Final Payable",
        "Payment Status",
    ]
    start_row = 7
    for col, title in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=title)
    _apply_header_style(ws, start_row, len(headers))

    money_cols = set(range(8, 17))
    for idx, emp in enumerate(payload.employees, start=start_row + 1):
        values = [
            emp.name,
            emp.employee_id,
            emp.department,
            emp.bank_name,
            emp.account_name,
            emp.account_number,
            emp.salary_month,
            _money_cell_value(emp.base_salary),
            _money_cell_value(emp.bonuses_total),
            _money_cell_value(emp.increments_total),
            _money_cell_value(emp.lateness_deduction),
            _money_cell_value(emp.absence_deduction),
            _money_cell_value(emp.early_sign_out_deduction),
            _money_cell_value(emp.manual_deductions),
            _money_cell_value(emp.other_adjustments),
            _money_cell_value(emp.total_deductions),
            _money_cell_value(emp.final_payable),
            emp.payment_status,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=val)
            if col in money_cols:
                cell.number_format = '#,##0.00'

    summary_row = start_row + len(payload.employees) + 2
    ws.cell(row=summary_row, column=1, value="Month Summary").font = Font(bold=True, size=12)
    summary_items = [
        ("Total Employees", payload.summary.employee_count),
        ("Total Base Salaries", _money_cell_value(payload.summary.total_base_salaries)),
        ("Total Bonuses", _money_cell_value(payload.summary.total_bonuses)),
        ("Total Salary Increments", _money_cell_value(payload.summary.total_increments)),
        ("Total Deductions", _money_cell_value(payload.summary.total_deductions)),
        ("Grand Total Payable", _money_cell_value(payload.summary.grand_total_payable)),
    ]
    for offset, (label, val) in enumerate(summary_items, start=1):
        ws.cell(row=summary_row + offset, column=1, value=label).font = Font(bold=True)
        val_cell = ws.cell(row=summary_row + offset, column=2, value=val)
        if isinstance(val, float):
            val_cell.number_format = '#,##0.00'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 24

    # --- Breakdown sheet ---
    bd = wb.create_sheet("Breakdowns")
    bd_headers = ["Employee", "Employee ID", "Category", "Description", "Date", "Amount (NGN)"]
    for col, title in enumerate(bd_headers, start=1):
        bd.cell(row=1, column=col, value=title)
    _apply_header_style(bd, 1, len(bd_headers))

    bd_row = 2
    for emp in payload.employees:
        sections: list[tuple[str, list[PayrollExportLineItem]]] = [
            ("Bonus", emp.bonus_lines),
            ("Salary Increment", emp.increment_lines),
            ("Late Coming", emp.lateness_lines),
            ("Absence", emp.absence_lines),
            ("Early Sign-Out", emp.early_sign_out_lines),
            ("Manual Deduction", emp.manual_deduction_lines),
        ]
        for category, lines in sections:
            for line in lines:
                bd.cell(row=bd_row, column=1, value=emp.name)
                bd.cell(row=bd_row, column=2, value=emp.employee_id)
                bd.cell(row=bd_row, column=3, value=category)
                if category in {"Bonus", "Salary Increment", "Manual Deduction"}:
                    bd.cell(row=bd_row, column=4, value=line.label)
                else:
                    bd.cell(row=bd_row, column=4, value=line.date_label or category)
                bd.cell(row=bd_row, column=5, value=line.date_label or "")
                amt_cell = bd.cell(row=bd_row, column=6, value=_money_cell_value(line.amount))
                amt_cell.number_format = '#,##0.00'
                bd_row += 1

    for col in range(1, len(bd_headers) + 1):
        bd.column_dimensions[get_column_letter(col)].width = 18
    bd.column_dimensions["A"].width = 24
    bd.column_dimensions["D"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def payload_to_api_dict(payload: PayrollExportPayload) -> dict:
    """Serialize for JSON API / PDF export page."""

    def _line(item: PayrollExportLineItem) -> dict:
        return {"label": item.label, "amount": str(item.amount), "date_label": item.date_label}

    def _emp(row: PayrollExportEmployeeRow) -> dict:
        return {
            "employee_id": row.employee_id,
            "name": row.name,
            "department": row.department,
            "bank_name": row.bank_name,
            "account_name": row.account_name,
            "account_number": row.account_number,
            "salary_month": row.salary_month,
            "base_salary": str(row.base_salary),
            "bonuses_total": str(row.bonuses_total),
            "increments_total": str(row.increments_total),
            "lateness_deduction": str(row.lateness_deduction),
            "absence_deduction": str(row.absence_deduction),
            "early_sign_out_deduction": str(row.early_sign_out_deduction),
            "manual_deductions": str(row.manual_deductions),
            "other_adjustments": str(row.other_adjustments),
            "total_deductions": str(row.total_deductions),
            "final_payable": str(row.final_payable),
            "payment_status": row.payment_status,
            "bonus_lines": [_line(x) for x in row.bonus_lines],
            "increment_lines": [_line(x) for x in row.increment_lines],
            "lateness_lines": [_line(x) for x in row.lateness_lines],
            "absence_lines": [_line(x) for x in row.absence_lines],
            "early_sign_out_lines": [_line(x) for x in row.early_sign_out_lines],
            "manual_deduction_lines": [_line(x) for x in row.manual_deduction_lines],
        }

    return {
        "company_name": payload.company_name,
        "payroll_month": payload.payroll_month,
        "payroll_status": payload.payroll_status,
        "generated_by": payload.generated_by,
        "generated_date": payload.generated_date,
        "generated_time": payload.generated_time,
        "period_id": payload.period_id,
        "period_year": payload.period_year,
        "period_month": payload.period_month,
        "employees": [_emp(e) for e in payload.employees],
        "summary": {
            "employee_count": payload.summary.employee_count,
            "total_base_salaries": str(payload.summary.total_base_salaries),
            "total_bonuses": str(payload.summary.total_bonuses),
            "total_increments": str(payload.summary.total_increments),
            "total_deductions": str(payload.summary.total_deductions),
            "grand_total_payable": str(payload.summary.grand_total_payable),
        },
    }
